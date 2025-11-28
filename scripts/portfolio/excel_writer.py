"""
ポートフォリオ自動生成システムのエクセル操作モジュール
テンプレートをコピーして書き込む方式を採用
"""

import logging
import os
import shutil

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
except ImportError as err:
    raise ImportError(
        "openpyxlがインストールされていません。"
        "以下のコマンドでインストールしてください: pip install openpyxl"
    ) from err

logger = logging.getLogger(__name__)


class ExcelWriteError(Exception):
    """エクセル書き込みエラー"""

    pass


class ExcelWriter:
    """エクセルファイルへの書き込みを管理するクラス"""

    def __init__(
        self, template_path: str, output_path: str, sheet_name: str = "フォーマット"
    ):
        """
        エクセルライターを初期化

        Args:
            template_path: テンプレートファイルのパス
            output_path: 出力ファイルのパス
            sheet_name: 操作するシート名

        Raises:
            ExcelWriteError: テンプレートファイルが見つからない場合
        """
        self.template_path = os.path.abspath(template_path)
        self.output_path = os.path.abspath(output_path)
        self.sheet_name = sheet_name
        self.workbook: Workbook | None = None
        self.sheet = None

        # テンプレートファイルの存在確認
        if not os.path.exists(self.template_path):
            raise ExcelWriteError(
                f"テンプレートファイルが見つかりません: {self.template_path}"
            )

        # テンプレートをコピー
        self._copy_template()

        # ワークブックを読み込む
        self._load_workbook()

    def _copy_template(self) -> None:
        """テンプレートファイルを出力パスにコピー"""
        try:
            # 出力ディレクトリが存在しない場合は作成
            output_dir = os.path.dirname(self.output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                logger.debug(f"出力ディレクトリを作成しました: {output_dir}")

            # テンプレートをコピー
            shutil.copy2(self.template_path, self.output_path)
            logger.info(
                f"テンプレートをコピーしました: {self.template_path} -> {self.output_path}"
            )

        except Exception as e:
            raise ExcelWriteError(f"テンプレートのコピーに失敗しました: {e}") from e

    def _load_workbook(self) -> None:
        """ワークブックを読み込む"""
        try:
            self.workbook = load_workbook(self.output_path)

            # シート名が存在するか確認（部分一致も許可）
            if self.sheet_name not in self.workbook.sheetnames:
                # 部分一致で探す
                matching_sheets = [
                    name
                    for name in self.workbook.sheetnames
                    if self.sheet_name.lower() in name.lower()
                ]
                if matching_sheets:
                    self.sheet_name = matching_sheets[0]
                    logger.warning(
                        f"シート名が完全一致しませんでした。部分一致を使用: {self.sheet_name}"
                    )
                else:
                    # 最初のシートを使用
                    self.sheet_name = self.workbook.sheetnames[0]
                    logger.warning(
                        f"指定されたシート名が見つかりませんでした。最初のシートを使用: {self.sheet_name}"
                    )

            self.sheet = self.workbook[self.sheet_name]
            logger.debug(f"ワークブックを読み込みました: {self.sheet_name}")

        except Exception as e:
            raise ExcelWriteError(f"ワークブックの読み込みに失敗しました: {e}") from e

    def write_cell(self, cell_address: str, value: str | int | float) -> None:
        """
        セルに値を書き込む（フォント色を黒に設定し、テキストの折り返しを有効化）

        Args:
            cell_address: セルアドレス（例: "D3", "A1"）
            value: 書き込む値
        """
        try:
            cell = self.sheet[cell_address]
            # 値を書き込む
            cell.value = value

            # フォント色を黒に設定
            current_font = cell.font
            black_color = "000000"  # 黒色（RGB形式）

            if current_font is None:
                cell.font = Font(color=black_color)
            else:
                # 既存のフォント設定を保持しつつ、色のみを黒に変更
                cell.font = Font(
                    name=current_font.name if current_font.name else "Calibri",
                    size=current_font.size if current_font.size else 11,
                    bold=current_font.bold,
                    italic=current_font.italic,
                    underline=current_font.underline,
                    strike=current_font.strike,
                    color=black_color,  # 黒色
                )

            # テキストの折り返しを有効化
            current_alignment = cell.alignment
            if current_alignment is None:
                cell.alignment = Alignment(wrap_text=True)
            else:
                # 既存の配置設定を保持しつつ、折り返しを有効化
                cell.alignment = Alignment(
                    horizontal=(
                        current_alignment.horizontal
                        if current_alignment.horizontal
                        else "general"
                    ),
                    vertical=(
                        current_alignment.vertical
                        if current_alignment.vertical
                        else "bottom"
                    ),
                    text_rotation=(
                        current_alignment.text_rotation
                        if current_alignment.text_rotation
                        else 0
                    ),
                    wrap_text=True,  # 折り返しを有効化
                    shrink_to_fit=(
                        current_alignment.shrink_to_fit
                        if hasattr(current_alignment, "shrink_to_fit")
                        else False
                    ),
                    indent=(
                        current_alignment.indent
                        if hasattr(current_alignment, "indent")
                        else 0
                    ),
                )

            logger.debug(f"セルに書き込みました: {cell_address} = {value}")

        except Exception as e:
            logger.error(f"セルへの書き込みに失敗しました: {cell_address}, エラー: {e}")
            raise ExcelWriteError(
                f"セルへの書き込みに失敗しました: {cell_address}"
            ) from e

    def write_cells(self, cell_data: dict[str, str | int | float]) -> None:
        """
        複数のセルに一括で値を書き込む

        Args:
            cell_data: セルアドレスをキー、値をバリューとする辞書
                      例: {"D3": "2025年11月26日", "D4": "タイトル"}
        """
        for cell_address, value in cell_data.items():
            try:
                self.write_cell(cell_address, value)
            except ExcelWriteError as e:
                logger.warning(
                    f"セル書き込みをスキップしました: {cell_address}, エラー: {e}"
                )
                # 続行（部分的な失敗を許容）

    def save(self) -> None:
        """ワークブックを保存"""
        try:
            if self.workbook is None:
                raise ExcelWriteError("ワークブックが読み込まれていません")

            self.workbook.save(self.output_path)
            logger.info(f"ワークブックを保存しました: {self.output_path}")

        except Exception as e:
            raise ExcelWriteError(f"ワークブックの保存に失敗しました: {e}") from e

    def close(self) -> None:
        """ワークブックを閉じる"""
        if self.workbook is not None:
            self.workbook.close()
            self.workbook = None
            self.sheet = None
            logger.debug("ワークブックを閉じました")

    def __enter__(self):
        """コンテキストマネージャーとして使用する場合のエントリ"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """コンテキストマネージャーとして使用する場合のエグジット"""
        if exc_type is None:
            # エラーがなければ保存
            self.save()
        self.close()
